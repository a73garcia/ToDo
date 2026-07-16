"""Aplicación local ToDo: servidor web, API REST y persistencia en Excel."""

from __future__ import annotations

import json
import mimetypes
import shutil
import threading
import time
import webbrowser
from datetime import datetime
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    APP_NAME,
    APP_VERSION,
    BACKUP_DIR,
    DEFAULT_PRIORITY,
    DEFAULT_STATE,
    EXCEL_COLUMNS,
    EXCEL_FILE,
    HOST,
    PORT,
    STATIC_DIR,
    TASK_PRIORITIES,
    TASK_STATES,
    TEMPLATE_DIR,
    ensure_directories,
)

TASK_FIELDS = [
    "id", "titulo", "descripcion", "responsable", "prioridad", "estado",
    "fecha_creacion", "fecha_inicio", "fecha_prevista", "fecha_finalizacion",
    "ultima_actualizacion", "avance", "observaciones",
]


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def ensure_workbook() -> None:
    """Crea un libro XLSX válido con las hojas necesarias."""
    ensure_directories()
    if EXCEL_FILE.exists():
        try:
            load_workbook(EXCEL_FILE).close()
            return
        except Exception:
            EXCEL_FILE.unlink(missing_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tareas"
    ws.append(EXCEL_COLUMNS)
    history = wb.create_sheet("Historial")
    history.append(["Fecha", "Tarea ID", "Acción", "Detalle"])
    config = wb.create_sheet("Configuración")
    config.append(["Clave", "Valor"])
    config.append(["Versión", APP_VERSION])
    config.append(["Estados", ", ".join(TASK_STATES)])
    config.append(["Prioridades", ", ".join(TASK_PRIORITIES)])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1976D2")
        sheet.freeze_panes = "A2"

    for index, width in enumerate([8, 30, 45, 24, 14, 16, 20, 16, 18, 20, 20, 12, 45], 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_FILE)


def backup_workbook() -> None:
    if not EXCEL_FILE.exists():
        return
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    shutil.copy2(EXCEL_FILE, BACKUP_DIR / f"tareas_{stamp}.xlsx")


def row_to_task(row: tuple) -> dict:
    values = list(row) + [None] * (len(TASK_FIELDS) - len(row))
    task = dict(zip(TASK_FIELDS, values[:len(TASK_FIELDS)]))
    task["id"] = int(task["id"]) if task["id"] not in (None, "") else None
    task["avance"] = int(task["avance"] or 0)
    for key, value in list(task.items()):
        if isinstance(value, datetime):
            task[key] = value.strftime("%Y-%m-%d %H:%M:%S")
        elif value is None:
            task[key] = "" if key not in ("id", "avance") else task[key]
    return task


def task_to_row(task: dict) -> list:
    return [task.get(field, "") for field in TASK_FIELDS]


class TaskStore:
    @staticmethod
    def all() -> list[dict]:
        ensure_workbook()
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Tareas"]
        tasks = [row_to_task(row) for row in ws.iter_rows(min_row=2, values_only=True) if row[0] is not None]
        wb.close()
        return tasks

    @staticmethod
    def get(task_id: int) -> dict | None:
        return next((task for task in TaskStore.all() if task["id"] == task_id), None)

    @staticmethod
    def create(data: dict) -> dict:
        ensure_workbook()
        backup_workbook()
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Tareas"]
        ids = [int(row[0]) for row in ws.iter_rows(min_row=2, values_only=True) if row[0] is not None]
        current = now_text()
        task = {
            "id": max(ids, default=0) + 1,
            "titulo": str(data.get("titulo", "")).strip(),
            "descripcion": str(data.get("descripcion", "")).strip(),
            "responsable": str(data.get("responsable", "")).strip(),
            "prioridad": data.get("prioridad") or DEFAULT_PRIORITY,
            "estado": data.get("estado") or DEFAULT_STATE,
            "fecha_creacion": current,
            "fecha_inicio": data.get("fecha_inicio", ""),
            "fecha_prevista": data.get("fecha_prevista") or data.get("fecha", ""),
            "fecha_finalizacion": "",
            "ultima_actualizacion": current,
            "avance": max(0, min(100, int(data.get("avance", 0) or 0))),
            "observaciones": str(data.get("observaciones", "")).strip(),
        }
        if not task["titulo"]:
            wb.close()
            raise ValueError("El título es obligatorio")
        if task["estado"] not in TASK_STATES:
            task["estado"] = DEFAULT_STATE
        if task["prioridad"] not in TASK_PRIORITIES:
            task["prioridad"] = DEFAULT_PRIORITY
        ws.append(task_to_row(task))
        TaskStore._history(wb, task["id"], "Creación", task["titulo"])
        wb.save(EXCEL_FILE)
        return task

    @staticmethod
    def update(task_id: int, data: dict) -> dict | None:
        ensure_workbook()
        backup_workbook()
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Tareas"]
        for row_number in range(2, ws.max_row + 1):
            if ws.cell(row_number, 1).value == task_id:
                task = row_to_task(tuple(ws.cell(row_number, column).value for column in range(1, len(TASK_FIELDS) + 1)))
                allowed = {
                    "titulo", "descripcion", "responsable", "prioridad", "estado",
                    "fecha_inicio", "fecha_prevista", "fecha_finalizacion", "avance", "observaciones",
                }
                changes = []
                for key in allowed:
                    if key in data:
                        value = data[key]
                        if key == "avance":
                            value = max(0, min(100, int(value or 0)))
                        if key == "estado" and value not in TASK_STATES:
                            continue
                        if key == "prioridad" and value not in TASK_PRIORITIES:
                            continue
                        if task.get(key) != value:
                            changes.append(f"{key}: {task.get(key)} → {value}")
                            task[key] = value
                if not str(task.get("titulo", "")).strip():
                    wb.close()
                    raise ValueError("El título es obligatorio")
                if task["estado"] == "Finalizada":
                    task["avance"] = 100
                    task["fecha_finalizacion"] = task.get("fecha_finalizacion") or now_text()
                elif "estado" in data and task["estado"] != "Finalizada":
                    task["fecha_finalizacion"] = ""
                task["ultima_actualizacion"] = now_text()
                for column, value in enumerate(task_to_row(task), 1):
                    ws.cell(row_number, column).value = value
                TaskStore._history(wb, task_id, "Actualización", "; ".join(changes) or "Sin cambios")
                wb.save(EXCEL_FILE)
                return task
        wb.close()
        return None

    @staticmethod
    def delete(task_id: int) -> bool:
        ensure_workbook()
        backup_workbook()
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Tareas"]
        for row_number in range(2, ws.max_row + 1):
            if ws.cell(row_number, 1).value == task_id:
                title = ws.cell(row_number, 2).value or ""
                ws.delete_rows(row_number)
                TaskStore._history(wb, task_id, "Eliminación", str(title))
                wb.save(EXCEL_FILE)
                return True
        wb.close()
        return False

    @staticmethod
    def dashboard() -> dict:
        tasks = TaskStore.all()
        today = datetime.now().strftime("%Y-%m-%d")
        result = {"total": len(tasks), "pendientes": 0, "en_curso": 0, "bloqueadas": 0,
                  "finalizadas": 0, "canceladas": 0, "retrasadas": 0, "avance_medio": 0}
        mapping = {"Pendiente": "pendientes", "En curso": "en_curso", "Bloqueada": "bloqueadas",
                   "Finalizada": "finalizadas", "Cancelada": "canceladas"}
        for task in tasks:
            key = mapping.get(task["estado"])
            if key:
                result[key] += 1
            due = str(task.get("fecha_prevista", ""))[:10]
            if due and due < today and task["estado"] not in ("Finalizada", "Cancelada"):
                result["retrasadas"] += 1
        if tasks:
            result["avance_medio"] = round(sum(int(t.get("avance", 0)) for t in tasks) / len(tasks), 1)
        return result

    @staticmethod
    def _history(wb, task_id: int, action: str, detail: str) -> None:
        wb["Historial"].append([now_text(), task_id, action, detail])


class TodoHandler(BaseHTTPRequestHandler):
    server_version = f"{APP_NAME}/{APP_VERSION}"

    def log_message(self, format_: str, *args) -> None:
        print(f"[{now_text()}] {format_ % args}")

    def send_json(self, payload, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", "0") or 0)
        if not length:
            return {}
        try:
            return json.loads(self.rfile.read(length).decode("utf-8"))
        except json.JSONDecodeError as exc:
            raise ValueError("JSON no válido") from exc

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        try:
            if path == "/api/health":
                return self.send_json({"ok": True, "app": APP_NAME, "version": APP_VERSION})
            if path == "/api/tasks":
                tasks = TaskStore.all()
                query = parse_qs(parsed.query).get("q", [""])[0].lower().strip()
                if query:
                    tasks = [t for t in tasks if query in " ".join(str(v) for v in t.values()).lower()]
                return self.send_json(tasks)
            if path.startswith("/api/tasks/"):
                task = TaskStore.get(self.task_id(path))
                return self.send_json(task, 200) if task else self.send_json({"error": "Tarea no encontrada"}, 404)
            if path == "/api/dashboard":
                return self.send_json(TaskStore.dashboard())
            if path == "/api/calendar":
                return self.send_json(TaskStore.all())
            if path == "/api/history":
                ensure_workbook()
                wb = load_workbook(EXCEL_FILE, read_only=True)
                rows = list(wb["Historial"].iter_rows(min_row=2, values_only=True))
                wb.close()
                return self.send_json([{"fecha": r[0], "task_id": r[1], "accion": r[2], "detalle": r[3]} for r in rows])
            return self.serve_static(path)
        except Exception as exc:
            return self.send_json({"error": str(exc)}, 500)

    def do_POST(self) -> None:
        try:
            if urlparse(self.path).path == "/api/tasks":
                return self.send_json(TaskStore.create(self.read_json()), 201)
            return self.send_json({"error": "Ruta no encontrada"}, 404)
        except ValueError as exc:
            return self.send_json({"error": str(exc)}, 400)
        except Exception as exc:
            return self.send_json({"error": str(exc)}, 500)

    def do_PATCH(self) -> None:
        self.handle_update()

    def do_PUT(self) -> None:
        self.handle_update()

    def handle_update(self) -> None:
        try:
            path = urlparse(self.path).path
            if not path.startswith("/api/tasks/"):
                return self.send_json({"error": "Ruta no encontrada"}, 404)
            task = TaskStore.update(self.task_id(path), self.read_json())
            return self.send_json(task) if task else self.send_json({"error": "Tarea no encontrada"}, 404)
        except ValueError as exc:
            return self.send_json({"error": str(exc)}, 400)
        except Exception as exc:
            return self.send_json({"error": str(exc)}, 500)

    def do_DELETE(self) -> None:
        try:
            path = urlparse(self.path).path
            if path.startswith("/api/tasks/") and TaskStore.delete(self.task_id(path)):
                return self.send_json({"ok": True})
            return self.send_json({"error": "Tarea no encontrada"}, 404)
        except Exception as exc:
            return self.send_json({"error": str(exc)}, 500)

    @staticmethod
    def task_id(path: str) -> int:
        try:
            return int(path.rstrip("/").split("/")[-1])
        except ValueError as exc:
            raise ValueError("ID de tarea no válido") from exc

    def serve_static(self, path: str) -> None:
        if path in ("", "/"):
            file_path = TEMPLATE_DIR / "index.html"
        elif path.startswith("/css/"):
            file_path = STATIC_DIR / "css" / path.removeprefix("/css/")
        elif path.startswith("/js/"):
            file_path = STATIC_DIR / "js" / path.removeprefix("/js/")
        else:
            file_path = TEMPLATE_DIR / path.lstrip("/")

        roots = (TEMPLATE_DIR.resolve(), STATIC_DIR.resolve())
        resolved = file_path.resolve()
        if not any(root == resolved or root in resolved.parents for root in roots) or not resolved.is_file():
            return self.send_json({"error": "Archivo no encontrado"}, 404)

        content = resolved.read_bytes()
        mime = mimetypes.guess_type(str(resolved))[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", f"{mime}; charset=utf-8" if mime.startswith("text/") else mime)
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)


def open_browser() -> None:
    time.sleep(1)
    webbrowser.open(f"http://{HOST}:{PORT}")


def main() -> None:
    ensure_workbook()
    print("=" * 60)
    print(f"{APP_NAME} {APP_VERSION}")
    print(f"Servidor: http://{HOST}:{PORT}")
    print(f"Excel: {EXCEL_FILE}")
    print("Pulse Ctrl+C para finalizar.")
    print("=" * 60)
    threading.Thread(target=open_browser, daemon=True).start()
    server = HTTPServer((HOST, PORT), TodoHandler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nCerrando servidor...")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
