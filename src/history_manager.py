"""
src/history_manager.py
Gestión del historial de cambios de las tareas.
"""

from datetime import datetime
from openpyxl import load_workbook

from config import EXCEL_FILE


class HistoryManager:
    SHEET_NAME = "Historial"

    def _workbook(self):
        return load_workbook(EXCEL_FILE)

    def add_entry(self, task_id, action, user="Sistema", observations=""):
        wb = self._workbook()

        if self.SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(self.SHEET_NAME)
            ws.append(["Fecha", "ID", "Acción", "Usuario", "Observaciones"])
        else:
            ws = wb[self.SHEET_NAME]

        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            task_id,
            action,
            user,
            observations
        ])

        wb.save(EXCEL_FILE)

    def get_history(self, task_id=None):
        wb = self._workbook()
        if self.SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[self.SHEET_NAME]
        result = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {
                "fecha": row[0],
                "task_id": row[1],
                "accion": row[2],
                "usuario": row[3],
                "observaciones": row[4],
            }
            if task_id is None or item["task_id"] == task_id:
                result.append(item)

        return result

    def clear_history(self):
        wb = self._workbook()
        if self.SHEET_NAME in wb.sheetnames:
            ws = wb[self.SHEET_NAME]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            wb.save(EXCEL_FILE)


if __name__ == "__main__":
    hm = HistoryManager()
    hm.add_entry(1, "Creación", "Sistema", "Registro de prueba")
    print(hm.get_history())
