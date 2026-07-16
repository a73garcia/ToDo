from pathlib import Path
from threading import RLock
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import EXCEL_FILE


class ExcelManager:
    TASKS_SHEET = "Tareas"
    HISTORY_SHEET = "Historial"
    SETTINGS_SHEET = "Configuración"

    TASK_COLUMNS = [
        "ID", "Título", "Descripción", "Responsable", "Prioridad", "Estado",
        "Fecha creación", "Fecha inicio", "Fecha prevista", "Fecha finalización",
        "Última actualización", "Avance (%)", "Observaciones", "Etiquetas",
        "Comentarios", "Favorito", "Recordatorio", "Proyecto",
        "Tiempo estimado", "Tiempo empleado", "Versión",
    ]
    HISTORY_COLUMNS = ["Fecha", "ID", "Acción", "Usuario", "Observaciones"]
    STRUCTURE_VERSION = 2
    _lock = RLock()

    def __init__(self, excel_path: Path = EXCEL_FILE):
        self.excel_path = Path(excel_path)

    def create_if_not_exists(self):
        with self._lock:
            if self.excel_path.exists():
                try:
                    workbook = load_workbook(self.excel_path)
                    workbook.close()
                    self.ensure_structure()
                    return False
                except Exception:
                    self.excel_path.unlink(missing_ok=True)

            self.excel_path.parent.mkdir(parents=True, exist_ok=True)
            workbook = Workbook()
            tasks = workbook.active
            tasks.title = self.TASKS_SHEET
            tasks.append(self.TASK_COLUMNS)

            history = workbook.create_sheet(self.HISTORY_SHEET)
            history.append(self.HISTORY_COLUMNS)

            settings = workbook.create_sheet(self.SETTINGS_SHEET)
            settings.append(["Propiedad", "Valor"])
            settings.append(["Versión de estructura", self.STRUCTURE_VERSION])
            settings.append(["Fecha de creación", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

            self._format(workbook)
            workbook.save(self.excel_path)
            workbook.close()
            return True

    def ensure_structure(self):
        if not self.excel_path.exists():
            return self.create_if_not_exists()

        with self._lock:
            workbook = load_workbook(self.excel_path)

            if self.TASKS_SHEET not in workbook.sheetnames:
                tasks = workbook.create_sheet(self.TASKS_SHEET, 0)
                tasks.append(self.TASK_COLUMNS)
            else:
                tasks = workbook[self.TASKS_SHEET]
                for index, header in enumerate(self.TASK_COLUMNS, 1):
                    tasks.cell(1, index, header)

            if self.HISTORY_SHEET not in workbook.sheetnames:
                history = workbook.create_sheet(self.HISTORY_SHEET)
                history.append(self.HISTORY_COLUMNS)
            else:
                history = workbook[self.HISTORY_SHEET]
                for index, header in enumerate(self.HISTORY_COLUMNS, 1):
                    history.cell(1, index, header)

            if self.SETTINGS_SHEET not in workbook.sheetnames:
                settings = workbook.create_sheet(self.SETTINGS_SHEET)
                settings.append(["Propiedad", "Valor"])
                settings.append(["Versión de estructura", self.STRUCTURE_VERSION])

            self._format(workbook)
            workbook.save(self.excel_path)
            workbook.close()
            return True

    def workbook(self, read_only=False, data_only=False):
        self.create_if_not_exists()
        return load_workbook(self.excel_path, read_only=read_only, data_only=data_only)

    def save(self, workbook):
        with self._lock:
            workbook.save(self.excel_path)

    @staticmethod
    def close(workbook):
        workbook.close()

    def next_id(self, worksheet):
        maximum = 0
        for (value,) in worksheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
            try:
                maximum = max(maximum, int(value))
            except (TypeError, ValueError):
                pass
        return maximum + 1

    def serialize_task(self, task):
        return task.to_excel_row()

    def append_history_row(self, row):
        with self._lock:
            workbook = self.workbook()
            worksheet = workbook[self.HISTORY_SHEET]
            worksheet.append(list(row))
            self.save(workbook)
            workbook.close()

    def _format(self, workbook):
        for sheet_name in (self.TASKS_SHEET, self.HISTORY_SHEET, self.SETTINGS_SHEET):
            if sheet_name not in workbook.sheetnames:
                continue
            worksheet = workbook[sheet_name]
            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        tasks = workbook[self.TASKS_SHEET]
        widths = [8, 28, 42, 22, 14, 16, 16, 16, 16, 18, 21, 12, 40, 28, 55, 12, 16, 22, 18, 18, 12]
        for index, width in enumerate(widths, 1):
            tasks.column_dimensions[get_column_letter(index)].width = width
