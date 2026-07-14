import csv
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from config import EXPORT_DIR


class ReportManager:

    def __init__(self, excel_manager, metrics_manager):

        self.excel = excel_manager
        self.metrics = metrics_manager

        self.export_folder = EXPORT_DIR
        self.export_folder.mkdir(
            parents=True,
            exist_ok=True
        )

    # =====================================================
    # RUTAS
    # =====================================================

    def excel_report(self):

        return self.export_folder / "Informe.xlsx"

    def csv_report(self):

        return self.export_folder / "Informe.csv"

    def pdf_report(self):

        return self.export_folder / "Informe.pdf"

    # =====================================================
    # EXPORTAR CSV
    # =====================================================

    def export_csv(self):

        fichero = self.csv_report()

        with open(
            fichero,
            "w",
            newline="",
            encoding="utf-8-sig"
        ) as f:

            writer = csv.writer(f)

            writer.writerow([

                "ID",

                "Título",

                "Responsable",

                "Estado",

                "Prioridad",

                "Categoría",

                "Inicio",

                "Prevista",

                "Avance"

            ])

            for tarea in self.excel.load_tasks():

                writer.writerow([

                    tarea.id,

                    tarea.titulo,

                    tarea.responsable,

                    tarea.estado,

                    tarea.prioridad,

                    tarea.categoria,

                    tarea.fecha_inicio,

                    tarea.fecha_prevista,

                    tarea.avance

                ])

        return fichero

    # =====================================================
    # EXPORTAR EXCEL
    # =====================================================

    def export_excel(self):

        wb = Workbook()

        ws = wb.active

        ws.title = "Informe"

        encabezado = [

            "ID",

            "Título",

            "Responsable",

            "Estado",

            "Prioridad",

            "Categoría",

            "Inicio",

            "Prevista",

            "Avance"

        ]

        ws.append(encabezado)

        azul = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        for celda in ws[1]:

            celda.fill = azul
            celda.font = Font(
                bold=True,
                color="FFFFFF"
            )

        for tarea in self.excel.load_tasks():

            ws.append([

                tarea.id,

                tarea.titulo,

                tarea.responsable,

                tarea.estado,

                tarea.prioridad,

                tarea.categoria,

                tarea.fecha_inicio,

                tarea.fecha_prevista,

                tarea.avance

            ])

        wb.save(
            self.excel_report()
        )

        return self.excel_report()

    # =====================================================
    # INFORME EJECUTIVO
    # =====================================================

    def executive_summary(self):

        return self.metrics.executive_summary()

    # =====================================================
    # EXPORTAR PDF
    # =====================================================

    def export_pdf(self):

        fichero = self.pdf_report()

        documento = SimpleDocTemplate(str(fichero))

        estilos = getSampleStyleSheet()

        elementos = []

        elementos.append(

            Paragraph(

                "<b>Task Planner Pro</b>",

                estilos["Title"]

            )

        )

        resumen = self.executive_summary()

        elementos.append(

            Paragraph(

                f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}",

                estilos["Normal"]

            )

        )

        elementos.append(

            Paragraph(

                "<br/>",

                estilos["Normal"]

            )

        )

        tabla = [

            ["Indicador", "Valor"],

            ["Total tareas", resumen["total_tareas"]],

            ["Tareas activas", resumen["tareas_activas"]],

            ["Finalizadas", resumen["tareas_finalizadas"]],

            ["Retrasadas", resumen["tareas_retrasadas"]],

            ["Avance global", f"{resumen['avance_global']} %"],

            ["Responsables", resumen["responsables"]],

            ["Categorías", resumen["categorias"]]

        ]

        t = Table(tabla)

        t.setStyle(

            TableStyle([

                ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E78")),

                ("TEXTCOLOR",(0,0),(-1,0),colors.white),

                ("GRID",(0,0),(-1,-1),0.5,colors.grey),

                ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

                ("ALIGN",(0,0),(-1,-1),"CENTER"),

                ("BOTTOMPADDING",(0,0),(-1,0),8),

            ])

        )

        elementos.append(t)

        documento.build(elementos)

        return fichero

    # =====================================================
    # INFORME DE PRODUCTIVIDAD
    # =====================================================

    def productivity_report(self):

        return self.metrics.owner_productivity()

    # =====================================================
    # INFORME POR RESPONSABLE
    # =====================================================

    def owner_report(self):

        datos = {}

        for tarea in self.excel.load_tasks():

            responsable = tarea.responsable or "Sin asignar"

            datos.setdefault(responsable, []).append(tarea)

        return datos

    # =====================================================
    # INFORME POR CATEGORÍA
    # =====================================================

    def category_report(self):

        datos = {}

        for tarea in self.excel.load_tasks():

            categoria = tarea.categoria or "Sin categoría"

            datos.setdefault(categoria, []).append(tarea)

        return datos

    # =====================================================
    # INFORME POR PRIORIDAD
    # =====================================================

    def priority_report(self):

        datos = {}

        for tarea in self.excel.load_tasks():

            prioridad = tarea.prioridad

            datos.setdefault(prioridad, []).append(tarea)

        return datos

    # =====================================================
    # MÉTRICAS
    # =====================================================

    def metrics_report(self):

        return self.metrics.full_report()



    # =====================================================
    # EXPORTAR TODO
    # =====================================================

    def export_all(self):

        return {

            "excel": self.export_excel(),

            "csv": self.export_csv(),

            "pdf": self.export_pdf()

        }

    # =====================================================
    # HISTORIAL DE INFORMES
    # =====================================================

    def report_history(self):

        historial = []

        for fichero in sorted(

            self.export_folder.iterdir(),

            key=lambda x: x.stat().st_mtime,

            reverse=True

        ):

            if not fichero.is_file():

                continue

            historial.append({

                "nombre": fichero.name,

                "fecha":

                    datetime.fromtimestamp(

                        fichero.stat().st_mtime

                    ),

                "tamano":

                    fichero.stat().st_size,

                "extension":

                    fichero.suffix

            })

        return historial

    # =====================================================
    # ELIMINAR INFORME
    # =====================================================

    def delete_report(self, fichero):

        fichero = Path(fichero)

        if fichero.exists():

            fichero.unlink()

            return True

        return False

    # =====================================================
    # LIMPIAR INFORMES
    # =====================================================

    def cleanup_reports(self, keep=30):

        informes = sorted(

            self.export_folder.glob("*.*"),

            key=lambda x: x.stat().st_mtime,

            reverse=True

        )

        eliminados = 0

        for fichero in informes[keep:]:

            try:

                fichero.unlink()

                eliminados += 1

            except Exception:

                pass

        return eliminados

    # =====================================================
    # TAMAÑO TOTAL
    # =====================================================

    def reports_size(self):

        total = 0

        for fichero in self.export_folder.glob("*.*"):

            total += fichero.stat().st_size

        return total

    # =====================================================
    # VALIDAR
    # =====================================================

    def validate(self):

        try:

            self.metrics.executive_summary()

            self.excel.load_tasks()

            return True

        except Exception:

            return False

    # =====================================================
    # INFORMACIÓN
    # =====================================================

    def info(self):

        return {

            "manager": "ReportManager",

            "version": "2.0",

            "export_folder":

                str(self.export_folder),

            "informes":

                len(

                    self.report_history()

                ),

            "tamano":

                self.reports_size(),

            "estado":

                self.validate()

        }

    # =====================================================
    # MANTENIMIENTO
    # =====================================================

    def maintenance(self):

        self.cleanup_reports()

    # =====================================================
    # INFORME COMPLETO
    # =====================================================

    def full_report(self):

        return {

            "executive":

                self.executive_summary(),

            "metrics":

                self.metrics_report(),

            "owners":

                self.owner_report(),

            "categories":

                self.category_report(),

            "priorities":

                self.priority_report(),

            "productivity":

                self.productivity_report(),

            "history":

                self.report_history()

        }









