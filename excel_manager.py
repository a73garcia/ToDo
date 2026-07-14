from openpyxl import load_workbook
from models import Task
from datetime import datetime
import shutil

    # =====================================================
    # ABRIR LIBRO
    # =====================================================

    def workbook(self):
        return load_workbook(self.file)

    # =====================================================
    # GUARDAR
    # =====================================================

    def save(self, wb):
        wb.save(self.file)

    # =====================================================
    # OBTENER SIGUIENTE ID
    # =====================================================

    def next_id(self):
        tareas = self.load_tasks()
        if not tareas:
            return 1
        return max(t.id for t in tareas) + 1

    # =====================================================
    # LEER TAREAS
    # =====================================================

    def load_tasks(self):

        wb = self.workbook()
        ws = wb[SHEET_TASKS]

        tareas = []

        for row in ws.iter_rows(min_row=2, values_only=True):

            if row[0] is None:
                continue

            t = Task()

            (
                t.id,
                t.titulo,
                t.descripcion,
                t.responsable,
                t.estado,
                t.prioridad,
                t.categoria,
                t.etiquetas,
                _proyecto,
                _horas_est,
                _horas_real,
                t.fecha_creacion,
                t.fecha_inicio,
                t.fecha_prevista,
                t.fecha_finalizacion,
                t.avance,
                _riesgo,
                t.comentarios,
            ) = row

            tareas.append(t)

        return tareas

    # =====================================================
    # INSERTAR
    # =====================================================

    def add_task(self, task):

        wb = self.workbook()
        ws = wb[SHEET_TASKS]

        ws.append([
            task.id,
            task.titulo,
            task.descripcion,
            task.responsable,
            task.estado,
            task.prioridad,
            task.categoria,
            task.etiquetas,
            "",
            "",
            "",
            task.fecha_creacion,
            task.fecha_inicio,
            task.fecha_prevista,
            task.fecha_finalizacion,
            task.avance,
            "",
            task.comentarios
        ])

        self.save(wb)

    # =====================================================
    # ACTUALIZAR
    # =====================================================

    def update_task(self, task):

        wb = self.workbook()
        ws = wb[SHEET_TASKS]

        for r in range(2, ws.max_row + 1):

            if ws.cell(r,1).value == task.id:

                ws.cell(r,2).value = task.titulo
                ws.cell(r,3).value = task.descripcion
                ws.cell(r,4).value = task.responsable
                ws.cell(r,5).value = task.estado
                ws.cell(r,6).value = task.prioridad
                ws.cell(r,7).value = task.categoria
                ws.cell(r,8).value = task.etiquetas
                ws.cell(r,12).value = task.fecha_creacion
                ws.cell(r,13).value = task.fecha_inicio
                ws.cell(r,14).value = task.fecha_prevista
                ws.cell(r,15).value = task.fecha_finalizacion
                ws.cell(r,16).value = task.avance
                ws.cell(r,18).value = task.comentarios
                break

        self.save(wb)

    # =====================================================
    # ELIMINAR TAREA
    # =====================================================

    def delete_task(self, task_id):

        wb = self.workbook()
        ws = wb[SHEET_TASKS]

        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == task_id:
                ws.delete_rows(row)
                break

        self.save(wb)

    # =====================================================
    # AÑADIR HISTORIAL
    # =====================================================

    def add_history(self, task_id, usuario, avance, comentario):

        wb = self.workbook()
        ws = wb[SHEET_HISTORY]

        ws.append([
            task_id,
            datetime.now().strftime(DATETIME_FORMAT),
            usuario,
            avance,
            comentario
        ])

        self.save(wb)

    # =====================================================
    # BACKUP
    # =====================================================

    def backup(self):

        if not AUTO_BACKUP:
            return

        destino = BACKUP_DIR / (
            "Backup_" +
            datetime.now().strftime("%Y%m%d_%H%M%S") +
            ".xlsx"
        )

        shutil.copy2(self.file, destino)

    # =====================================================
    # ESTADÍSTICAS
    # =====================================================

    def statistics(self):

        tareas = self.load_tasks()

        datos = {
            "Pendiente": 0,
            "En curso": 0,
            "Bloqueada": 0,
            "Finalizada": 0,
            "Total": len(tareas)
        }

        for t in tareas:
            if t.estado in datos:
                datos[t.estado] += 1

        return datos

    # =====================================================
    # ACTUALIZAR MÉTRICAS POWER BI
    # =====================================================

    def update_metrics(self):

        wb = self.workbook()
        ws = wb["Metricas"]

        datos = self.statistics()

        total = datos["Total"]

        avance = 0

        if total:
            suma = sum(t.avance for t in self.load_tasks())
            avance = round(suma / total)

        ws.append([
            datetime.now().strftime(DATE_FORMAT),
            datos["Pendiente"],
            datos["En curso"],
            datos["Bloqueada"],
            datos["Finalizada"],
            0,
            total,
            avance
        ])

        self.save(wb)

    # =====================================================
    # REFRESCO COMPLETO
    # =====================================================

    def refresh(self):

        self.update_metrics()

        self.backup()



